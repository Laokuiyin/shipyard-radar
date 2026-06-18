from datetime import date, datetime

from openpyxl import load_workbook

from shipwatch.db import Database
from shipwatch.domain import Article, Extraction, Milestone
from shipwatch.exporter import ExcelExporter
from shipwatch.merge import ProjectMerger


def test_export_has_required_sheets_and_source_url(tmp_path):
    db = Database(tmp_path / "test.db")
    db.init()
    article_id = db.upsert_article(
        Article(
            source_id="jiangnan",
            yard_hint="江南造船",
            channel="官网",
            title="新船开工",
            url="https://example.com/news/1",
            content="江南造船为某航运建造一艘集装箱船开工",
            published_at=date(2026, 1, 2),
            fetched_at=datetime.now(),
            content_hash="abc",
        )
    )
    extraction = Extraction(
        relevant=True,
        yard="江南造船",
        owner_project="某航运",
        ship_type="集装箱船",
        ship_count=1,
        current_progress="开工",
        milestones=[Milestone("start", "开工", date(2026, 1, 2), evidence="正式开工")],
        confidence=0.9,
    )
    db.mark_extracted(article_id, extraction)
    ProjectMerger(db).merge(article_id, extraction)
    output = tmp_path / "out.xlsx"
    ExcelExporter(db).export(output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "项目主表", "今日新增与变更", "项目里程碑", "来源明细",
        "待人工复核", "来源采集状态",
    ]
    assert workbook["项目主表"]["J2"].value == "https://example.com/news/1"
    assert workbook["项目主表"]["J2"].hyperlink.target == "https://example.com/news/1"


def test_failure_reason_is_normalized():
    assert ExcelExporter._normalize_failure_reason(
        "搜狗链接未跳转到微信原文: https://weixin.sogou.com/antispider/"
    ) == "搜狗反爬；原文跳转失败"
