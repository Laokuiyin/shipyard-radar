from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from shipwatch.db import Database


HEADERS = {
    "项目主表": [
        "船厂", "船东/项目", "船型", "船数", "当前项目进度", "开工时间",
        "预计/实际完工时间", "其他关键节点", "消息标题", "消息来源（网址）",
        "消息发布日期", "采集时间", "置信度", "复核状态",
    ],
    "今日新增与变更": [
        "变更时间", "船厂", "船东/项目", "船型", "船数", "当前项目进度",
        "消息标题", "消息来源（网址）", "复核状态",
    ],
    "项目里程碑": [
        "船厂", "船东/项目", "船型", "节点", "节点日期", "是否预计",
        "原文证据", "消息标题", "消息来源（网址）",
    ],
    "来源明细": [
        "船厂提示", "渠道", "公众号", "消息标题", "消息来源（网址）",
        "消息发布日期", "采集时间", "抓取状态", "错误/备注", "是否相关",
    ],
    "待人工复核": [
        "船厂", "船东/项目", "船型", "船数", "当前进度", "复核原因",
        "置信度", "消息标题", "消息来源（网址）",
    ],
    "来源采集状态": [
        "船厂/来源主体", "来源渠道", "采集状态", "最后执行时间",
        "最后成功时间", "发现数量", "失败原因",
    ],
}


class ExcelExporter:
    def __init__(
        self,
        db: Database,
        source_names: dict[str, str] | None = None,
        source_channels: dict[str, list[tuple[str, str]]] | None = None,
    ):
        self.db = db
        self.source_names = source_names or {}
        self.source_channels = source_channels or {}

    def export(self, output_path: Path, changed_since: str | None = None) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheets = {name: workbook.create_sheet(name) for name in HEADERS}
        for name, headers in HEADERS.items():
            sheets[name].append(headers)

        projects = self.db.query(
            """
            SELECT p.*,
              (SELECT GROUP_CONCAT(label || COALESCE('：' || event_date, ''), '；')
               FROM milestones m WHERE m.project_id=p.id) AS milestones_text,
              a.title AS source_title, a.url AS source_url,
              a.published_at, a.fetched_at
            FROM projects p
            LEFT JOIN project_sources ps ON ps.project_id=p.id
            LEFT JOIN articles a ON a.id=ps.article_id
            WHERE ps.article_id=(
              SELECT MAX(ps2.article_id) FROM project_sources ps2 WHERE ps2.project_id=p.id
            )
            ORDER BY p.last_changed_at DESC
            """
        )
        for row in projects:
            sheets["项目主表"].append(
                [
                    row["yard"], row["owner_project"], row["ship_type"], row["ship_count"],
                    row["current_progress"], row["start_date"], row["completion_date"],
                    row["milestones_text"], row["source_title"], row["source_url"],
                    row["published_at"], self._display_datetime(row["fetched_at"]),
                    row["confidence"], row["review_status"],
                ]
            )
            if not changed_since or row["last_changed_at"] >= changed_since:
                sheets["今日新增与变更"].append(
                    [
                        self._display_datetime(row["last_changed_at"]), row["yard"],
                        row["owner_project"], row["ship_type"],
                        row["ship_count"], row["current_progress"], row["source_title"],
                        row["source_url"], row["review_status"],
                    ]
                )
            if row["review_status"] == "待复核":
                sheets["待人工复核"].append(
                    [
                        row["yard"], row["owner_project"], row["ship_type"], row["ship_count"],
                        row["current_progress"], row["review_reason"], row["confidence"],
                        row["source_title"], row["source_url"],
                    ]
                )

        milestones = self.db.query(
            """
            SELECT p.yard, p.owner_project, p.ship_type, m.label, m.event_date,
                   m.is_expected, m.evidence, a.title, a.url
            FROM milestones m JOIN projects p ON p.id=m.project_id
            JOIN articles a ON a.id=m.article_id
            ORDER BY COALESCE(m.event_date, a.published_at), m.id
            """
        )
        for row in milestones:
            sheets["项目里程碑"].append(
                [
                    row["yard"], row["owner_project"], row["ship_type"], row["label"],
                    row["event_date"], "是" if row["is_expected"] else "否", row["evidence"],
                    row["title"], row["url"],
                ]
            )

        articles = self.db.query("SELECT * FROM articles ORDER BY fetched_at DESC")
        for row in articles:
            sheets["来源明细"].append(
                [
                    row["yard_hint"], row["channel"], row["account_name"], row["title"], row["url"],
                    row["published_at"], self._display_datetime(row["fetched_at"]),
                    row["fetch_status"], row["fetch_error"],
                    "是" if row["relevant"] else ("否" if row["relevant"] == 0 else "待处理"),
                ]
            )

        states = {
            row["source_key"]: row
            for row in self.db.query("SELECT * FROM crawl_state ORDER BY source_key")
        }
        article_health = {
            (row["source_id"], row["channel"]): row
            for row in self.db.query(
                """
                SELECT source_id, channel, COUNT(*) AS total,
                       SUM(CASE WHEN fetch_status='ok' THEN 1 ELSE 0 END) AS ok_count,
                       SUM(CASE WHEN fetch_status!='ok' THEN 1 ELSE 0 END) AS failed_count,
                       GROUP_CONCAT(DISTINCT fetch_error) AS errors
                FROM articles GROUP BY source_id, channel
                """
            )
        }
        source_ids = list(self.source_names) or sorted(
            {key.split(":", 1)[0] for key in states}
        )
        for source_id in source_ids:
            channels = self.source_channels.get(
                source_id, [("website", "官网"), ("wechat", "微信公众号")]
            )
            for channel_key, channel_name in channels:
                key = f"{source_id}:{channel_key}"
                row = states.get(key)
                if not row:
                    values = [
                        self.source_names.get(source_id, source_id), channel_name, "未执行",
                        None, None, None, "尚无采集记录",
                    ]
                else:
                    if row["last_error"] and row["result_count"]:
                        status = "发现成功 / 正文获取失败"
                    else:
                        status = "失败" if row["last_error"] else "成功"
                    health = article_health.get((source_id, channel_name))
                    article_error = None
                    if not row["last_error"] and health and health["failed_count"]:
                        if not health["ok_count"]:
                            status = "发现成功 / 正文获取失败"
                        else:
                            status = "部分失败"
                        article_error = health["errors"]
                    values = [
                        self.source_names.get(source_id, source_id),
                        channel_name,
                        status,
                        self._display_datetime(row["last_attempt_at"]),
                        self._display_datetime(row["last_success_at"]),
                        row["result_count"],
                        self._normalize_failure_reason(row["last_error"] or article_error),
                    ]
                sheets["来源采集状态"].append(values)

        for name, sheet in sheets.items():
            self._style_sheet(sheet, name)
        workbook.save(output_path)
        return output_path

    @staticmethod
    def _style_sheet(sheet, name: str) -> None:
        header_fill = PatternFill("solid", fgColor="17365D")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 28
        widths = {
            "A": 18, "B": 25, "C": 28, "D": 10, "E": 18, "F": 15, "G": 20,
            "H": 42, "I": 38, "J": 55, "K": 16, "L": 21, "M": 12, "N": 12,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")
        if sheet.max_row > 1:
            ref = f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
            table = Table(displayName=f"T_{abs(hash(name))}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)

    @staticmethod
    def _display_datetime(value: str | None) -> str | None:
        return value[:19].replace("T", " ") if value else None

    @staticmethod
    def _normalize_failure_reason(value: str | None) -> str | None:
        if not value:
            return None
        reasons = []
        checks = (
            ("antispider", "搜狗反爬"),
            ("反爬", "搜狗反爬"),
            ("验证码", "微信/搜狗验证码"),
            ("appmsgcaptcha", "微信验证码"),
            ("未跳转到微信原文", "原文跳转失败"),
            ("正文过短", "正文为空或过短"),
            ("SSL", "官网 TLS/SSL 连接失败"),
            ("TLS", "官网 TLS/SSL 连接失败"),
            ("待补采", "正文尚未成功获取"),
        )
        for keyword, label in checks:
            if keyword.lower() in value.lower() and label not in reasons:
                reasons.append(label)
        return "；".join(reasons) if reasons else value[:500]


def daily_output_path(output_dir: Path, day: date | None = None) -> Path:
    target = day or date.today()
    return output_dir / f"船厂新船项目_{target.isoformat()}.xlsx"
